"""
엑셀 파일을 불러옵니다
"""


import os


from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill


def load_excel(excel_file_path=str) -> Workbook:
    """엑셀 파일을 불러옵니다. 파일이 존재하지 않은 경우 HEADER를 추가하고 새 파일을 불러옵니다"""
    if not os.path.exists(excel_file_path):
        wb = Workbook()
        ws = wb.active

        headers = ['위원회', '피감기관', '위원', 'BOOK_ID',
                   'SEQNO', 'level', '질의', 'PDF상 답변', 'REALFILE_NAME', '파일명', '실제경로', 'FILE_NAME']
        header_color = PatternFill(start_color='4f81bd',
                                   end_color='4f81bd', fill_type='solid')

        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
            ws.cell(row=1, column=col_idx).fill = header_color
        wb.save(excel_file_path)
    wb = load_workbook(excel_file_path)

    return wb
